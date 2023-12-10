# -*- coding: utf-8 -*-
import pickle
from pathlib import Path
import streamlit_authenticator as stauth
import streamlit as st

import datetime
import PIL.Image
from streamlit_option_menu import option_menu
from datetime import date
from streamlit_autorefresh import st_autorefresh
import openpyxl
from spire.xls import *
from spire.xls.common import *
import webbrowser
from datetime import datetime, time  # Import the 'time' class

st.set_page_config(page_title='Πίνακας',
                       page_icon='logo.png',
                       layout='wide', initial_sidebar_state='expanded')
st.experimental_set_query_params(show_map=True, selected=["111hfffffjhjhfddddssjjhjhjhjhjjhjhfxdxszsszaasia",
                                                              "dfdfdfdsiopfjhh6556h56jhx;s5665656sdsisahhhhpppppp655655656dfdfdfhghghjhamerica"], )
st_autorefresh(60000)
names = ["PeterParker", "RebeccaMiller"]
usernames = ["pparker", "rmiller"]

file_path = Path(__file__).parent/"hashed_pw.pkl"
with file_path.open("rb") as file:
    hashed_passwords = pickle.load(file)
authenticator = stauth.Authenticate(names, usernames, hashed_passwords, "ekloges9","algdef", cookie_expiry_days=2)
name, authentication_status, username = authenticator.login("Login", "main")

if authentication_status ==False:
    st.error("Username/password is incorrect")
if authentication_status == None:
    st.warning("Please enter username and password")
if authentication_status:

    with st.sidebar.container():
        image = PIL.Image.open("logo.png")
        st.image(image, use_column_width=True)

    with st.sidebar:
        selected = option_menu(
            menu_title="Μενού",
            options=["Αναχωρήσεις","Δρομολόγια"],
            icons=["house", "house"],
            menu_icon="cast",
            default_index=0,
            # orientation="horizontal",
            styles={
                "container": {"padding": "!important", "backgroundcolor": "grey"},
                "icon": {"color": "green", "font-size": "18px"},
                "nav-link": {
                    "font-size": "18px",
                    "text-align": "left",
                    "margin": "0px",
                    "--hover-color": "#DFD8D8",
                },
                "nav-link-selected": {"background-color": "#464545"}
            },
        )


    st.sidebar.markdown("Για να δείτε όλα τα δρομολόγια ΚΤΕΛ μπείτε στο https://www.ktenimathias/dromologia")
    st.sidebar.markdown('''Created & Designed with ❤️ by [scax.gr](https://www.scax.gr).''')
    hide_st_style = """
                <style>
                #MainMenu {visibility: hidden;}
                footer {visibility: hidden;}
                header {visibility: hidden;}
                </style>
                """
    st.markdown(hide_st_style, unsafe_allow_html=True)

    if selected == "Αναχωρήσεις":
        # using now() to get current time
        # Printing attributes of now().
        current_time = datetime.now()
        #print("Year : ", current_time.year)
        #print("Month : ", current_time.month)
        #print("Day : ", current_time.day)
        #print("Hour : ", current_time.hour)
        #print("Minute : ", current_time.minute)
        print("-------------------------------")
        #print("UTC Time: ", datetime.datetime.utcnow())
        today = date.today()
        if today.weekday() == 0:
            hmera= "Monday"
            hmeran= 0
        elif today.weekday() == 1:
            hmera = "Tuesday"
            hmeran=1
        elif today.weekday() == 2:
            hmera = "Wednesday"
            hmeran=2
        elif today.weekday() == 3:
            hmera = "Thursday"
            hmeran=3
        elif today.weekday() == 4:
            hmera = "Friday"
            hmeran=4
        elif today.weekday() == 5:
            hmera = "Saturday"
            hmeran=5
        elif today.weekday() == 6:
            hmera = "Sunday"
            hmeran=6
        print(f"Today is:{hmera}, {hmeran}")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
                    <style>
                    .big-font {
                        font-size:50px !important;
                        text-align: left;
                    }
                    </style>
                    """, unsafe_allow_html=True)

            st.markdown('<p class="big-font">Αναχωρήσεις</p>', unsafe_allow_html=True)
        with col2:
            today = date.today()
            # dd/mm/YY
            d1 = today.strftime("**%d/%m/%Y**")
            st.write('')
            st.write('**ΗΜΕΡΟΜΗΝΙΑ:**', d1)
            now = datetime.now()
            current_time = now.strftime("**%H:%M:%S**")
            st.write('**ΤΕΛΕΥΤΑΙΑ ΕΝΗΜΕΡΩΣΗ ΑΝΑΧΩΡΗΣΕΩΝ:**', current_time)
        # Specify the path to your Excel file
        excel_file_path = 'Dromologia.xlsx'

        # Open the workbook
        workbook = openpyxl.load_workbook(excel_file_path)

        # Get the sheet names
        sheet_count = len(workbook.sheetnames)
        print(f"Sheet count: {sheet_count}")
        for sheet_name in workbook.sheetnames:
            print(f"Sheet name: {sheet_name}")
            sheet = workbook[sheet_name]
            if hmeran <= 4:
                if sheet_name.startswith('Κ'):


                    # Specify the row and column indices (replace these with your actual indices)
                    column_letter = 'B'  # Change this to the column letter you're interested in

                    # Access the sheet
                    sheet = workbook[sheet_name]

                    # Get the current time
                    current_time = datetime.now().time()

                    # Iterate through the cells in the specified column
                    for cell in sheet[column_letter]:
                        # Check if the cell contains a time value (float representing time)
                        if isinstance(cell.value, (float,datetime,  time)):
                            # Convert float or datetime to time if needed
                            cell_time = (cell.value.time()if isinstance(cell.value, (float, datetime))else cell.value)

                            # Check if the time in the cell is greater than or equal to the current time
                            if cell_time >= current_time:
                                # Print the cell value and stop the loop
                                print(f'Cell value: {cell.value}, Column: {column_letter}, Row: {cell.row}')
                                # Specify the row and column indices (replace these with your actual indices)
                                row_index = 2
                                column_index = 1
                                cell_proorismos_value = sheet.cell(row=row_index, column=column_index).value
                                st.write(f"## {cell_proorismos_value} : {cell.value}")
                                break
            if hmeran > 4 & hmeran<=6:
                if sheet_name.startswith('Σ'):

                    # Specify the row and column indices (replace these with your actual indices)
                    column_letter = 'B'  # Change this to the column letter you're interested in

                    # Access the sheet
                    sheet = workbook[sheet_name]

                    # Get the current time
                    current_time = datetime.now().time()

                    # Iterate through the cells in the specified column
                    for cell in sheet[column_letter]:
                        # Check if the cell contains a time value (float representing time)
                        if isinstance(cell.value, (float, datetime, time)):
                            # Convert float or datetime to time if needed
                            cell_time = (
                                cell.value.time() if isinstance(cell.value, (float, datetime)) else cell.value)

                            # Check if the time in the cell is greater than or equal to the current time
                            if cell_time >= current_time:
                                # Print the cell value and stop the loop
                                print(f'Cell value: {cell.value}, Column: {column_letter}, Row: {cell.row}')
                                # Specify the row and column indices (replace these with your actual indices)
                                row_index = 2
                                column_index = 1
                                cell_proorismos_value = sheet.cell(row=row_index, column=column_index).value
                                st.write(f"## {cell_proorismos_value} : {cell.value}")
                                break

                    # Close the workbook
                    workbook.close()




    if selected == "Δρομολόγια":
        webbrowser.open('https://ktelnimathias.gr/', new=2, autoraise=True)
     #pass
